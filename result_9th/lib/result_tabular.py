import json

import openpyxl

from lib import db
from lib.calculatedmarks import calculate
from lib.codehelper import get_column_config_for_subject, \
    get_safe_output_xls_path, cfg, is_red_high_req
from lib.excelhelper import Cell, all_borders, add_table, \
    save_close_and_start, apply_column_widths
from lib.uihelper import MyApp

d = {}


def calculate_marks():
    for k, v in d['marksMap'].items():
        calculate(v, 'all', ())


def load_data():
    d['studentMap'] = db.get_student_map(d['ddDivision'].get())
    d['marksMap'] = db.get_marks_map_for_all_subjects(d['ddDivision'].get())
    d['colInfo'] = get_column_config_for_subject(d['examnm'])
    calculate_marks()


def get_output_file_path():
    return get_safe_output_xls_path(f"{d['ddDivision'].get()}_{d['examnm']}",
                                    True)


def add_excel_base_columns(wb):
    sht = wb.active
    Cell(7, 1, sht).set("अ. नं.")
    Cell(7, 2, sht).set("हजेरी क्रमांक").wrap()
    sht.cell(7, 3).value = "विद्यार्थ्याचे नाव"
    sht.cell(9, 3).value = "ठेवलेले गुण"
    sht.column_dimensions["A"].width = 8
    sht.column_dimensions["B"].width = 8
    sht.column_dimensions["C"].width = 30
    sht.row_dimensions[7].height = 120
    all_borders(sht, 1, 7, 3, 9)


def add_excel_header(wb: openpyxl.Workbook):
    sht = wb.active
    ar = [[cfg['school name']],
          [],
          ["परीक्षा", d['examnm']],
          ["तुकडी", d['ddDivision'].get()]]
    add_table(sht, 2, 2, ar)


def add_excel_student_info(wb):
    sht = wb.active
    for i, v in enumerate(d['studentMap']):
        sid, roll, nm = v
        Cell(10 + i, 1, sht).set(i + 1).border()
        Cell(10 + i, 2, sht).set(roll).border()
        Cell(10 + i, 3, sht).set(nm).wrap().border()


def excel_format_mark_cells(wb):
    sht = wb.active
    for r in range(len(d['studentMap'])):
        for c in range(len(d['colInfo'])):
            Cell(10 + r, 4 + c, sht).border()


def add_excel_exam_info(wb):
    sht = wb.active
    for i, examcol in enumerate(d['colInfo']):
        nm = examcol['nm']
        total = examcol['total']
        alias = examcol.get('alias', '')
        Cell(7, 4 + i, sht).set(nm).border().verticalwrap()
        Cell(8, 4 + i, sht).set(alias).border()
        Cell(9, 4 + i, sht).set(total).border()
        width = 6 if "width" not in examcol else examcol["width"]
        apply_column_widths(sht, 4 + i, [width])


def add_excel_marks(wb):
    sht = wb.active
    for ir, student in enumerate(d['studentMap']):
        for ic, exam in enumerate(d['colInfo']):
            if student[0] in d['marksMap'] \
                    and exam['id'] in d['marksMap'][student[0]]:
                c = Cell(10 + ir, 4 + ic, sht)
                c.set(d['marksMap'][student[0]][exam['id']]).wrap()
                if is_red_high_req(d['marksMap'][student[0]], exam):
                    c.red_color()

def populate_excel(wb):
    add_excel_header(wb)
    add_excel_base_columns(wb)
    load_data()
    add_excel_student_info(wb)
    add_excel_exam_info(wb)
    excel_format_mark_cells(wb)
    add_excel_marks(wb)


def generate(examnm):
    d['examnm'] = examnm
    filepath = get_output_file_path()
    wb = openpyxl.Workbook()
    populate_excel(wb)
    save_close_and_start(wb, filepath)


def show_ui(app: MyApp, examnm: str):
    d['app'] = app
    app.clear_screen()
    frm_select = app.main_frame.add_frame("Select", 500, 160, [1, 1, 1, 1])
    frm_select.add_label("Division", "Division", 18, 1, [2, 1, 1, 1])
    d['ddDivision'] = frm_select.add_dropdown("ddDivision",
                                              db.get_division_list(),
                                              25, 1, [2, 2, 1, 1], lambda x: 1)
    frm_select.add_button("btnGenerate", "Generate",
                          lambda: generate(examnm), [3, 2, 1, 2])
