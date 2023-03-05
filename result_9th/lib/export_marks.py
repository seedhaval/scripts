import openpyxl
from openpyxl.utils import get_column_letter

from lib import db
from lib.calculatedmarks import calculate
from lib.codehelper import get_safe_output_xls_path, cfg, \
    get_column_config_for_subject
from lib.excelhelper import Cell, all_borders, add_table, save_close_and_start
from lib.uihelper import MyApp

d = {}


def calculate_marks():
    cols = [x['id'] for x in d['colInfo'] if x['type'] == 'calculated']
    for k, v in d['marksMap'].items():
        calculate(v, 'subset', cols)


def load_data():
    d['studentMap'] = db.get_student_map(d['ddDivision'].get())
    d['examMap'] = db.get_exam_map(d['ddSubject'].get())
    d['marksMap'] = db.get_marks_map(d['ddSubject'].get(),
                                     d['ddDivision'].get())
    d['colInfo'] = get_column_config_for_subject(d['ddSubject'].get())
    calculate_marks()


def get_output_file_path():
    return get_safe_output_xls_path(f"{d['ddDivision'].get()}_"
                                    f"{d['ddSubject'].get()}", True)


def add_excel_base_columns(wb):
    sht = wb.active
    Cell(7, 1, sht).set("अ. नं.")
    Cell(7, 2, sht).set("हजेरी क्रमांक").wrap()
    sht.cell(7, 3).value = "विद्यार्थ्याचे नाव"
    sht.cell(9, 3).value = "ठेवलेले गुण"
    sht.column_dimensions["A"].width = 8
    sht.column_dimensions["B"].width = 8
    sht.column_dimensions["C"].width = 30
    sht.row_dimensions[7].height = 120
    all_borders(sht, 1, 7, 3, 9)


def add_excel_header(wb: openpyxl.Workbook):
    sht = wb.active
    ar = [[cfg['school name']],
          [],
          ["विषय", d['ddSubject'].get()],
          ["तुकडी", d['ddDivision'].get()]]
    add_table(sht, 2, 2, ar)


def add_excel_student_info(wb):
    sht = wb.active
    for i, v in enumerate(d['studentMap']):
        sid, roll, nm = v
        Cell(10 + i, 1, sht).set(i + 1).border()
        Cell(10 + i, 2, sht).set(roll).border()
        Cell(10 + i, 3, sht).set(nm).wrap().border()


def excel_format_mark_cells(wb):
    sht = wb.active
    for r in range(len(d['studentMap'])):
        for c in range(len(d['colInfo'])):
            Cell(10 + r, 4 + c, sht).border().color(d['colInfo'][c]['color'])


def add_excel_exam_info(wb):
    sht = wb.active
    for i, examcol in enumerate(d['colInfo']):
        if 'nm' in examcol:
            nm = examcol['nm']
            total = examcol['total']
        else:
            nm = d['examMap'][int(examcol['id'])][0]
            total = d['examMap'][int(examcol['id'])][1]
        alias = examcol.get('alias', '')
        clr = examcol['color']
        Cell(7, 4 + i, sht).set(nm).border().verticalwrap().color(clr)
        Cell(8, 4 + i, sht).set(alias).border().color(clr)
        Cell(9, 4 + i, sht).set(total).border().color(clr)
        sht.column_dimensions[get_column_letter(4 + i)].width = 6


def add_excel_marks(wb):
    sht = wb.active
    for ir, student in enumerate(d['studentMap']):
        for ic, exam in enumerate(d['colInfo']):
            if student[0] in d['marksMap'] \
                    and exam['id'] in d['marksMap'][student[0]]:
                Cell(10 + ir, 4 + ic, sht).set(
                    d['marksMap'][student[0]][exam['id']])


def populate_excel(wb):
    add_excel_header(wb)
    add_excel_base_columns(wb)
    load_data()
    add_excel_student_info(wb)
    add_excel_exam_info(wb)
    excel_format_mark_cells(wb)
    add_excel_marks(wb)


def export_data():
    filepath = get_output_file_path()
    wb = openpyxl.Workbook()
    populate_excel(wb)
    save_close_and_start(wb, filepath)


def show_ui(app: MyApp):
    d['app'] = app
    app.clear_screen()
    frm_select = app.main_frame.add_frame("Select", 500, 160, [1, 1, 1, 1])
    frm_select.add_label("Subject", "Subject", 18, 1, [1, 1, 1, 1])
    d['ddSubject'] = frm_select.add_dropdown("ddSubject",
                                             db.get_subject_list(),
                                             25, 1, [1, 2, 1, 1], lambda x: 1)
    frm_select.add_label("Division", "Division", 18, 1, [2, 1, 1, 1])
    d['ddDivision'] = frm_select.add_dropdown("ddDivision",
                                              db.get_division_list(),
                                              25, 1, [2, 2, 1, 1], lambda x: 1)
    frm_select.add_button("btnExport", "Export", export_data, [3, 2, 1, 2])
