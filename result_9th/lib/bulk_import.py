from tkinter import filedialog, messagebox

import openpyxl
from openpyxl.utils import get_column_letter

from lib import excelhelper, db
from lib.codehelper import output_path, get_safe_output_xls_path, cfg
from lib.excelhelper import Cell, protect_sheet, save_close_and_start, \
    add_table, all_borders
from lib.uihelper import MyApp

d = {}


def load_data():
    d['studentMap'] = db.get_student_map(d['ddDivision'].get())
    d['examMap'] = db.get_exam_map(d['ddSubject'].get())
    d['marksMap'] = db.get_marks_map(d['ddSubject'].get(),
                                     d['ddDivision'].get())


def get_output_file_path():
    return get_safe_output_xls_path(
        f"{d['ddDivision'].get()}_{d['ddSubject'].get()}_bulk_template"
        , False)


def add_excel_base_columns(wb):
    sht = wb.active
    Cell(7, 1, sht).set("sid")
    Cell(7, 2, sht).set("हजेरी क्रमांक").wrap()
    sht.cell(7, 3).value = "विद्यार्थ्याचे नाव"
    sht.cell(9, 3).value = "ठेवलेले गुण"
    sht.column_dimensions["A"].width = 8
    sht.column_dimensions["B"].width = 8
    sht.column_dimensions["C"].width = 30
    sht.row_dimensions[7].height = 120
    all_borders(sht, 1, 7, 3, 9)


def add_excel_header(wb: openpyxl.Workbook):
    sht = wb.active
    ar = [[cfg['school name']],
          [],
          ["विषय", d['ddSubject'].get()],
          ["तुकडी", d['ddDivision'].get()]]
    add_table(sht, 2, 2, ar)


def add_excel_student_info(wb):
    sht = wb.active
    for i, v in enumerate(d['studentMap']):
        sid, roll, nm = v
        Cell(10 + i, 1, sht).set(sid).border()
        Cell(10 + i, 2, sht).set(roll).border()
        Cell(10 + i, 3, sht).set(nm).wrap().border()


def excel_format_mark_cells(wb):
    sht = wb.active
    for r in range(len(d['studentMap'])):
        for c in range(len(d['examMap'])):
            Cell(10 + r, 4 + c, sht).border().unprotect()


def add_excel_exam_info(wb):
    sht = wb.active
    for i, examid in enumerate(sorted(d['examMap'].keys())):
        nm = d['examMap'][examid][0]
        total = d['examMap'][examid][1]
        Cell(7, 4 + i, sht).set(nm).border().verticalwrap()
        Cell(8, 4 + i, sht).set(examid).border()
        Cell(9, 4 + i, sht).set(total).border()
        sht.column_dimensions[get_column_letter(4 + i)].width = 6


def add_excel_marks(wb):
    sht = wb.active
    for ir, student in enumerate(d['studentMap']):
        for ic, examid in enumerate(sorted(d['examMap'].keys())):
            if student[0] in d['marksMap'] \
                    and str(examid) in d['marksMap'][student[0]]:
                Cell(10 + ir, 4 + ic, sht).set(
                    d['marksMap'][student[0]][str(examid)])


def populate_excel(wb):
    add_excel_header(wb)
    add_excel_base_columns(wb)
    load_data()
    add_excel_student_info(wb)
    add_excel_exam_info(wb)
    excel_format_mark_cells(wb)
    add_excel_marks(wb)


def export_data():
    filepath = get_output_file_path()
    wb = openpyxl.Workbook()
    populate_excel(wb)
    protect_sheet(wb)
    save_close_and_start(wb, filepath)


def import_data():
    file = filedialog.askopenfilename(initialdir=output_path)
    if not file:
        return
    data = excelhelper.read_all_rows(file, False)
    subject = data[3][2]
    division = data[4][2]
    examids = data[7][3:]
    studentids = [x[0] for x in data[9:]]
    if subject != d['ddSubject'].get() or division != d['ddDivision'].get():
        messagebox.showerror("Invalid file", "Invalid file")
        return
    out = []
    for ir, row in enumerate(data[9:]):
        for ic, col in enumerate(row[3:]):
            if str(col).strip() not in ('None', ''):
                out.append(f"({studentids[ir]}, {examids[ic]}, {col})")
    db.delete_marks_for_div_subject(d['ddSubject'].get(), d['ddDivision'].get())
    db.bulk_insert_marks(out)
    messagebox.showinfo("Done !!", "Done !!")


def show_ui(app: MyApp):
    d['app'] = app
    app.clear_screen()
    frm_select = app.main_frame.add_frame("Select", 500, 160, [1, 1, 1, 1])
    frm_select.add_label("Subject", "Subject", 18, 1, [1, 1, 1, 1])
    d['ddSubject'] = frm_select.add_dropdown("ddSubject", db.get_subject_list(),
                                             25, 1, [1, 2, 1, 1], lambda x: 1)
    frm_select.add_label("Division", "Division", 18, 1, [2, 1, 1, 1])
    d['ddDivision'] = frm_select.add_dropdown("ddDivision",
                                              db.get_division_list(),
                                              25, 1, [2, 2, 1, 1], lambda x: 1)
    frm_select.add_button("btnGenTemplate", "Generate Template", export_data,
                          [3, 1, 1, 1])
    frm_select.add_button("btnImport", "Import", import_data,
                          [3, 2, 1, 1])
