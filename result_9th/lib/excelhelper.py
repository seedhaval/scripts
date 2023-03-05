import os
import random
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import get_column_letter as cl
from openpyxl.worksheet.pagebreak import Break

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


class Cell:
    def __init__(self, row, col, sht):
        self.cell = sht.cell(row, col)
        self.sht = sht
        self.row = row
        self.col = col

    def wrap(self):
        self.cell.alignment = Alignment(wrap_text=True)
        return self

    def verticalwrap(self):
        self.cell.alignment = Alignment(wrap_text=True, textRotation=90)
        return self

    def center(self):
        self.cell.alignment = Alignment(horizontal='center')
        return self

    def color(self, clr):
        self.cell.font = Font(color=clr)
        return self

    def set(self, v):
        self.cell.value = v
        return self

    def border(self):
        self.cell.border = thin_border
        return self

    def unprotect(self):
        self.cell.protection = Protection(locked=False)


def read_all_rows(fl, shtnm):
    out = []
    wb = openpyxl.load_workbook(fl)
    if shtnm:
        sht = wb[shtnm]
    else:
        sht = wb.worksheets[0]
    for r in range(sht.max_row):
        row = []
        for c in range(sht.max_column):
            row.append(sht.cell(r + 1, c + 1).value)
        out.append(row)
    wb.close()
    return out


def protect_sheet(wb):
    sht = wb.active
    sht.protection.password = str(random.randint(10000, 99999))


def save_close_and_start(wb, filepath):
    wb.save(filepath)
    wb.close()
    os.startfile(filepath)


def add_table(sht, left, top, ar):
    for ir, row in enumerate(ar):
        for ic, col in enumerate(row):
            sht.cell(top + ir, left + ic).value = col


def all_borders(sht, left, top, right, bottom):
    for ir in range(top, bottom + 1):
        for ic in range(left, right + 1):
            sht.cell(ir, ic).border = thin_border


def apply_column_widths(sht, left, ar):
    for ic, val in enumerate(ar):
        sht.column_dimensions[cl(ic + left)].width = val


def load_template(filepath, shtnm, left, top, right, bottom):
    d = {}
    wb = openpyxl.load_workbook(filepath)
    sht = wb[shtnm]
    d['widths'] = [sht.column_dimensions[cl(x)].width for x in
                   range(left, right + 1)]
    # left, top, right, bottom
    ar = [x.bounds for x in sht.merged_cells.ranges]
    shift_ar = []
    for row in ar:
        l, t, r, b = row
        if l >= left and r <= right and t >= top and b <= bottom:
            shift_ar.append([l - left, t - top, r - left, b - top])
    d['merged'] = shift_ar
    out = []
    for ir in range(top, bottom + 1):
        row = []
        for ic in range(left, right + 1):
            cell = sht.cell(ir, ic)
            row.append({
                "value": cell.value
                , "font": copy(cell.font)
                , "border": copy(cell.border)
                , "fill": copy(cell.fill)
                , "number_format": copy(cell.number_format)
                , "protection": copy(cell.protection)
                , "alignment": copy(cell.alignment)
            })
        out.append(row)
    wb.close()
    d['data'] = out
    return d


def apply_template(sht, tmplt, left, top):
    for ir, row in enumerate(tmplt['data']):
        for ic, col in enumerate(row):
            cell = sht.cell(ir + top, ic + left)
            cell.value = col["value"]
            cell.font = col["font"]
            cell.border = col["border"]
            cell.fill = col["fill"]
            cell.number_format = col["number_format"]
            cell.protection = col["protection"]
            cell.alignment = col["alignment"]

    for minfo in tmplt['merged']:
        l, t, r, b = minfo
        range = f"{cl(l + left)}{t + top}:{cl(r + left)}{b + top}"
        sht.merge_cells(range)


def add_page_break(sht, rowid):
    page_break = Break(id=rowid)
    sht.row_breaks.append(page_break)
