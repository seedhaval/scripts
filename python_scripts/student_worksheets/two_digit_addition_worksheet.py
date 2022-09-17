from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import os
import random

num_problems = 24
problem_cols_per_page = 3
problem_rows_per_page = 4
problem_page_break_at = problem_cols_per_page * problem_rows_per_page
regular = Side(border_style='medium', color='000000')
wb = Workbook()
ws = wb.active


def write_sum(left: int, top: int, num1: int, num2: int, ws):
    t_o_num1 = (' ' + str(num1))[-2:]
    t_o_num2 = (' ' + str(num2))[-2:]
    ws.cell(top, left + 1).value = 'T'
    ws.cell(top, left + 2).value = 'O'
    ws.cell(top + 2, left + 1).value = int(t_o_num1[0])
    ws.cell(top + 2, left + 2).value = int(t_o_num1[1])
    if t_o_num2[0] != ' ':
        ws.cell(top + 3, left + 1).value = int(t_o_num2[0])
    ws.cell(top + 3, left + 2).value = int(t_o_num2[1])
    ws.cell(top + 3, left).value = '+'
    for row in range(top, top + 5):
        for col in range(left, left + 3):
            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row, col).font = Font(size=20)
            ws.cell(row, col).border = Border(top=regular, bottom=regular, left=regular, right=regular)


for i in range(1, 26):
    ws.column_dimensions[get_column_letter(i)].width = 6

for i in range(num_problems):
    left = ((i%problem_cols_per_page)*5) + 1
    top = ((i//problem_cols_per_page)*7) + 1
    num1 = random.randint(10,98)
    num2 = random.randint(1,(99-num1))
    write_sum(left, top, num1, num2, ws)
    if( i+1) % problem_page_break_at == 0:
        print('here')
        print(top)
        ws.row_breaks.append(Break(id=top+5))

wb.save('out.xlsx')
os.startfile('out.xlsx')
