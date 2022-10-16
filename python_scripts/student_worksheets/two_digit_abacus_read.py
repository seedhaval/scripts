from PIL import Image, ImageDraw, ImageFont
import openpyxl.drawing.image
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import os
import random

num_problems = 10
cols_per_page = 2
rows_per_page = 4
font = ImageFont.truetype("arial.ttf", 14)
problem_page_break_at = cols_per_page * rows_per_page
regular = Side(border_style='medium', color='000000')
wb = Workbook()
ws = wb.active

problem_width_cells = 7
problem_height_cells = 12
bean_radius = 7


def draw_circles(xcenter, bottom, count, radius, draw):
    btm = bottom
    for i in range(count):
        top = btm - (2 * radius)
        draw.ellipse((xcenter - radius, top, xcenter + radius, btm),
                     fill=(255, 255, 255, 255), outline=(0, 0, 0, 255))
        btm = top


def add_image(left: int, top: int, num: int, ws, idx: int):
    img = Image.new("RGBA", (290, 220))
    draw = ImageDraw.Draw(img)
    draw.rectangle((0, 0, 286, 216), outline=(0, 0, 0, 255))
    draw.rectangle((225, 90, 275, 126), outline=(0, 0, 0, 255))
    draw.line((20, 180, 100, 180), fill=(0, 0, 0, 255))
    draw.line((120, 180, 200, 180), fill=(0, 0, 0, 255))
    draw.text((40, 190), "Tens", font=font, fill=(0, 0, 0, 255))
    draw.text((140, 190), "Ones", font=font, fill=(0, 0, 0, 255))
    tens = num // 10
    ones = num % 10
    draw_circles(60, 180, tens, bean_radius, draw)
    draw_circles(160, 180, ones, bean_radius, draw)
    draw.line((60, 20, 60, 180 - (tens * 2 * bean_radius)), fill=(0, 0, 0, 255))
    draw.line((160, 20, 160, 180 - (ones * 2 * bean_radius)), fill=(0, 0, 0,
                                                                    255))
    img.save('tmp' + str(idx) + '.png')
    ws_img = openpyxl.drawing.image.Image('tmp' + str(idx) + '.png')
    ws_img.anchor = get_column_letter(left) + str(top)
    ws.add_image(ws_img)


def write_problem(left: int, top: int, num: int, ws, idx: int):
    add_image(left, top, num, ws, idx)


for i in range(1, 26):
    ws.column_dimensions[get_column_letter(i)].width = 6

ar = list(range(1, 100))
random.shuffle(ar)
for i in range(num_problems):
    left = ((i % cols_per_page) * problem_width_cells) + 1
    top = ((i // cols_per_page) * problem_height_cells) + 1
    num = random.randint(10, 98)
    write_problem(left, top, ar[i], ws, i)
    if (i + 1) % problem_page_break_at == 0:
        ws.row_breaks.append(Break(id=top + problem_height_cells - 1))

wb.save('out.xlsx')
for i in range(num_problems):
    os.remove('tmp' + str(i) + '.png')
os.startfile('out.xlsx')
