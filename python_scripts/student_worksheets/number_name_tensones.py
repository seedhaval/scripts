from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import os
import random
from num2words import num2words

num_problems = 20
problem_page_break_at = 20
regular = Side(border_style='medium', color='000000')
wb = Workbook()
ws = wb.active

ar = list(range(1, 100))
random.shuffle(ar)


def gettensones(num):
    return f'{num // 10} tens + {num % 10} ones'


def getname(num):
    return num2words(num).replace('-',' ')


words_per_typ = num_problems // 3
out = []
out.extend([(str(x), '', '') for x in ar[:words_per_typ]])
out.extend(
    [('', gettensones(x), '') for x in ar[words_per_typ:words_per_typ * 2]])
out.extend([('', '', getname(x)) for x in ar[words_per_typ * 2:num_problems]])
random.shuffle(out)


def write_problem(left: int, top: int, data, ws):
    ws.row_dimensions[top].height = 30
    for i, v in enumerate(data):
        ws.cell(top, left + i).value = v
        ws.cell(top, left + i).font = Font(size=14)
        ws.cell(top, left + i).alignment = Alignment(wrap_text=True,
                                                     vertical='center',indent=1)
        ws.cell(top, left + i).border = Border(left=regular, right=regular,
                                               top=regular, bottom=regular)

def add_header(left,top,ws):
    data = ('Number','Tens and Ones form','Number Name')
    write_problem(left,top,data,ws)

widths = ((1, 12), (2, 40), (3, 30))
for c, w in widths:
    ws.column_dimensions[get_column_letter(c)].width = w

i = 0
top = 1
left = 1
add_header(left,top,ws)
top += 1
while i < num_problems:
    data = out[i]
    write_problem(left, top, data, ws)
    i += 1
    if i % problem_page_break_at == 0:
        ws.row_breaks.append(Break(id=top))
        top += 1
        add_header(left, top, ws)
    top += 1

wb.save('out.xlsx')
os.startfile('out.xlsx')
