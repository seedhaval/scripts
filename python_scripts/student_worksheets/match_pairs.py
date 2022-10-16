from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import os
import random

num_problems = 20
sections_per_page = 3
problems_per_section = 5

problem_page_break_at = problems_per_section * sections_per_page
regular = Side(border_style='medium', color='000000')
wb = Workbook()
ws = wb.active

with open("match.txt") as f:
    mtch = [x.split() for x in f.readlines() if x.strip()]

out = []
while len(out) < num_problems:
    random.shuffle(mtch)
    out.extend(mtch)


def get_section(strt, end):
    d = {'left': [x[0] for x in out[strt:end]]}
    d['right'] = [x[1] for x in out[strt:end]]
    random.shuffle(d['right'])
    return d


def write_problem(left: int, top: int, data, ws):
    print(data)
    for i in range(len(data['left'])):
        ws.cell(top + i, left).value = data['left'][i]
        ws.cell(top + i, left).font = Font(size=14)
        ws.cell(top + i, left).alignment = Alignment(wrap_text=True,
                                                     vertical='center')
        ws.cell(top + i, left + 2).value = data['right'][i]
        ws.cell(top + i, left + 2).font = Font(size=14)
        ws.cell(top + i, left + 2).alignment = Alignment(wrap_text=True,
                                                         vertical='center')
        ws.row_dimensions[top + i].height = 30
    for i in range(-1, 4):
        ws.cell(top + len(data['left']) + 1, left + i).border = Border(
            bottom=regular)


widths = ((1, 4), (2, 35), (3, 5), (4, 35))
for c, w in widths:
    ws.column_dimensions[get_column_letter(c)].width = w

i = 0
top = 2
while i < num_problems:
    left = 2
    data = get_section(i, i + problems_per_section)
    write_problem(left, top, data, ws)
    i += problems_per_section
    if i % problem_page_break_at == 0:
        ws.row_breaks.append(Break(id=top + problems_per_section +2))
    top += problems_per_section + 3

wb.save('out.xlsx')
os.startfile('out.xlsx')
